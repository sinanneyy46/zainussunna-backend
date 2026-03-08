"""
Views for Academic Admission System
RESTful API with state machine enforcement and schema generation.
"""
import logging
from rest_framework import viewsets, status, views, serializers
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny, IsAdminUser
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db import transaction
from django.db.models import Count, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
from io import BytesIO
import urllib.parse

# Django cache for performance
from django.core.cache import cache

# Get logger for this module
logger = logging.getLogger(__name__)

from .models import (
    Program, ProgramField, Admission, AdmissionState,
    AdmissionStateLog, AdmissionEvent, InternalNote,
    ContentPage, Achievement, GalleryItem, Enquiry, AnalyticEvent, Faculty,
    WhatsAppConfig
)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# WeasyPrint for PDF generation
try:
    from weasyprint import HTML
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False


# ==============================================================================
# SECTION 1: DASHBOARD OVERVIEW API
# ==============================================================================

class DashboardView(views.APIView):
    """
    Dashboard Overview API - High-level statistics in a single response.
    Optimized with minimal queries using aggregation.
    """
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Try to get from cache first (60 seconds)
        cache_key = 'admission_dashboard_stats'
        cached_data = cache.get(cache_key)
        
        if cached_data:
            return Response(cached_data)
        
        # Get all admissions with select_related for program
        admissions = Admission.objects.select_related('program').all()
        
        # Total count - single query
        total_applications = admissions.count()
        
        # State distribution - single aggregated query
        state_counts = dict(admissions.values('state').annotate(
            count=Count('id')
        ).values_list('state', 'count'))
        
        # Program distribution - single aggregated query
        program_counts = dict(admissions.values('program__name').annotate(
            count=Count('id')
        ).values_list('program__name', 'count'))
        
        # Build response
        response_data = {
            'total_applications': total_applications,
            'draft': state_counts.get('draft', 0),
            'submitted': state_counts.get('submitted', 0),
            'under_review': state_counts.get('under_review', 0),
            'approved': state_counts.get('approved', 0),
            'rejected': state_counts.get('rejected', 0),
            'program_distribution': {
                'integrated_shareea': program_counts.get('Integrated Shareea', 0),
                'thahfeel_quran': program_counts.get('Thahfeel-ul Quran', 0),
            },
            'admission_funnel': {
                'draft': state_counts.get('draft', 0),
                'submitted': state_counts.get('submitted', 0),
                'under_review': state_counts.get('under_review', 0),
                'approved': state_counts.get('approved', 0),
                'rejected': state_counts.get('rejected', 0),
            },
            'pending_reviews': state_counts.get('submitted', 0) + state_counts.get('under_review', 0),
        }
        
        # Cache for 60 seconds
        cache.set(cache_key, response_data, 60)
        
        return Response(response_data)


# ==============================================================================
# SECTION 2: ADMISSIONS CONTROL PANEL
# ==============================================================================

class AdmissionControlPanelSerializer(serializers.Serializer):
    """SECTION 2: Serializer for control panel with action flags"""
    id = serializers.UUIDField()
    application_number = serializers.CharField()
    program_name = serializers.CharField()
    name = serializers.CharField()
    phone = serializers.CharField()
    district = serializers.CharField()
    state = serializers.CharField()
    submitted_at = serializers.DateTimeField(allow_null=True)
    can_review = serializers.SerializerMethodField()
    can_approve = serializers.SerializerMethodField()
    can_reject = serializers.SerializerMethodField()
    is_possible_duplicate = serializers.SerializerMethodField()
    
    def get_can_review(self, obj):
        return obj.state == AdmissionState.SUBMITTED
    
    def get_can_approve(self, obj):
        return obj.state == AdmissionState.UNDER_REVIEW
    
    def get_can_reject(self, obj):
        return obj.state == AdmissionState.UNDER_REVIEW
    
    def get_is_possible_duplicate(self, obj):
        # Check for duplicates based on phone/email
        phone_exists = Admission.objects.filter(phone=obj.phone).exclude(id=obj.id).exists()
        if phone_exists:
            return True
        
        if obj.email:
            email_exists = Admission.objects.filter(email=obj.email).exclude(id=obj.id).exists()
            if email_exists:
                return True
        
        return False


class AdmissionViewSet(viewsets.ModelViewSet):
    """
    Admission API - State machine controlled.
    Frontend cannot skip, jump, or override steps.
    
    Public (unauthenticated) users can CREATE new admissions.
    Authenticated users can list, retrieve, update, and manage admissions.
    """
    queryset = Admission.objects.select_related('program').all()
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    
    def get_permissions(self):
        """Allow public (unauthenticated) users to create and submit admissions."""
        if self.action in ['create', 'complete_step', 'submit', 'status']:
            return [AllowAny()]
        return [IsAuthenticated()]
    
    def get_serializer_class(self):
        from .serializers import (
            AdmissionListSerializer, AdmissionDetailSerializer,
            AdmissionCreateSerializer, AdmissionStepSerializer,
            AdmissionSubmitSerializer, StateTransitionSerializer
        )
        
        if self.action == 'list':
            return AdmissionListSerializer
        elif self.action == 'create':
            return AdmissionCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return AdmissionStepSerializer
        return AdmissionDetailSerializer
    
    def get_queryset(self):
        """Filter admissions based on user role with enhanced filtering"""
        queryset = Admission.objects.select_related('program').all()
        
        # SECTION 2: Enhanced filtering
        state = self.request.query_params.get('state')
        if state:
            queryset = queryset.filter(state=state)
        
        program = self.request.query_params.get('program')
        if program:
            queryset = queryset.filter(
                Q(program__slug=program) | Q(program__name__icontains=program)
            )
        
        district = self.request.query_params.get('district')
        if district:
            queryset = queryset.filter(address_district__icontains=district)
        
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            queryset = queryset.filter(created_at__gte=date_from)
        if date_to:
            queryset = queryset.filter(created_at__lte=date_to)
        
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(application_number__icontains=search) |
                Q(phone__icontains=search) |
                Q(email__icontains=search)
            )
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Enhanced list with action flags"""
        queryset = self.get_queryset()
        
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = AdmissionControlPanelSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = AdmissionControlPanelSerializer(queryset, many=True)
        return Response(serializer.data)
    
    def create(self, request, *args, **kwargs):
        """Create new admission with step 1 data"""
        from .serializers import AdmissionCreateSerializer, AdmissionDetailSerializer
        
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            admission = serializer.save()
            
            logger.info(
                f"Admission created: {admission.application_number}, "
                f"program={admission.program.name if admission.program else 'N/A'}, "
                f"step={admission.current_step}"
            )
            
            duplicate_warning = self._check_duplicates(admission)
            
            output_serializer = AdmissionDetailSerializer(admission)
            response_data = output_serializer.data
            if duplicate_warning:
                response_data['duplicate_warning'] = duplicate_warning
            
            return Response(response_data, status=status.HTTP_201_CREATED)
        except Exception as e:
            import traceback
            print(f"Error creating admission: {e}")
            traceback.print_exc()
            return Response(
                {'error': str(e), 'detail': 'Failed to create admission'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    def _check_duplicates(self, admission):
        """SECTION 8: Check for potential duplicates"""
        duplicates = []
        
        phone_matches = Admission.objects.filter(
            phone=admission.phone
        ).exclude(id=admission.id).values_list('application_number', flat=True)
        
        if phone_matches:
            duplicates.append({
                'type': 'phone',
                'matches': list(phone_matches)
            })
        
        if admission.email:
            email_matches = Admission.objects.filter(
                email=admission.email
            ).exclude(id=admission.id).values_list('application_number', flat=True)
            
            if email_matches:
                duplicates.append({
                    'type': 'email',
                    'matches': list(email_matches)
                })
        
        if admission.photo_hash:
            photo_matches = Admission.objects.filter(
                photo_hash=admission.photo_hash
            ).exclude(id=admission.id).values_list('application_number', flat=True)
            
            if photo_matches:
                duplicates.append({
                    'type': 'photo',
                    'matches': list(photo_matches)
                })
        
        return duplicates if duplicates else None
    
    @action(detail=True, methods=['get'])
    def timeline(self, request, pk=None):
        """SECTION 3: Get admission timeline"""
        admission = self.get_object()
        
        timeline = []
        
        timeline.append({
            'event': 'application_created',
            'time': admission.created_at.isoformat() if admission.created_at else None,
            'display': 'Application Created'
        })
        
        events = AdmissionEvent.objects.filter(
            admission=admission,
            event_type='step_completed'
        ).order_by('created_at')
        
        for event in events:
            step = event.event_data.get('step', 1)
            timeline.append({
                'event': 'step_completed',
                'step': step,
                'time': event.created_at.isoformat() if event.created_at else None,
                'display': f'Step {step} Completed'
            })
        
        if admission.submitted_at:
            timeline.append({
                'event': 'submitted',
                'time': admission.submitted_at.isoformat(),
                'display': 'Application Submitted'
            })
        
        logs = AdmissionStateLog.objects.filter(
            admission=admission
        ).order_by('created_at')
        
        for log in logs:
            if log.new_state != 'draft':
                timeline.append({
                    'event': 'state_changed',
                    'state': log.new_state,
                    'time': log.created_at.isoformat() if log.created_at else None,
                    'display': f'Status: {log.new_state.replace("_", " ").title()}'
                })
        
        notes_count = admission.notes.count()
        if notes_count > 0:
            timeline.append({
                'event': 'notes_added',
                'count': notes_count,
                'display': f'{notes_count} Internal Note(s)'
            })
        
        return Response(timeline)
    
    @action(detail=True, methods=['post'])
    def complete_step(self, request, pk=None):
        """Complete current step and advance"""
        from .serializers import AdmissionStepSerializer
        
        admission = self.get_object()
        
        logger.debug(
            f"complete_step: admission {admission.id}, "
            f"current_step={admission.current_step}, "
            f"completed_steps={admission.completed_steps}"
        )
        
        if admission.state != AdmissionState.DRAFT:
            logger.warning(
                f"Step completion blocked: admission {admission.application_number} "
                f"is not in draft state (state={admission.state})"
            )
            return Response(
                {'error': 'Cannot modify submitted admission'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        completed_steps = admission.completed_steps if isinstance(admission.completed_steps, list) else []
        expected_next_step = 1
        for s in [1, 2, 3]:
            if s not in completed_steps:
                expected_next_step = s
                break
        
        if admission.current_step != expected_next_step:
            logger.warning(
                f"Step bypass attempt blocked: admission {admission.application_number}, "
                f"expected step {expected_next_step}, got {admission.current_step}"
            )
            return Response(
                {'error': f'Cannot complete step {admission.current_step}. Please complete steps in order.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = AdmissionStepSerializer(admission, data=request.data)
        
        print(f"DEBUG complete_step: request.data = {request.data}")
        
        if not serializer.is_valid():
            print(f"DEBUG complete_step: validation errors = {serializer.errors}")
            return Response(
                {'error': 'Validation failed', 'details': serializer.errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            admission = Admission.objects.select_for_update().get(pk=admission.pk)
            serializer.save()
            
            logger.info(
                f"Step completed: admission {admission.application_number}, "
                f"step={admission.current_step - 1}, "
                f"completed_steps={admission.completed_steps}"
            )
        
        from .serializers import AdmissionDetailSerializer
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data)
    
    @action(detail=True, methods=['post'])
    def submit(self, request, pk=None):
        """Submit the admission - final submission."""
        from .serializers import AdmissionSubmitSerializer, AdmissionDetailSerializer
        
        logger.info(f"Submission attempt for admission pk={pk}")
        
        admission = self.get_object()
        
        if admission.state != AdmissionState.DRAFT:
            logger.warning(
                f"Submission blocked: admission {admission.application_number} "
                f"already submitted (state={admission.state})"
            )
            return Response(
                {"error": "Admission already submitted"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        completed_steps = admission.completed_steps if isinstance(admission.completed_steps, list) else []
        required_steps = {1, 2, 3}
        completed_set = set(completed_steps)
        
        if completed_set != required_steps:
            logger.warning(
                f"Submission blocked: admission {admission.application_number} "
                f"has not completed all steps (completed={completed_steps})"
            )
            return Response(
                {"error": "All steps must be completed before submission"},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        serializer = AdmissionSubmitSerializer(admission, data={})
        if not serializer.is_valid():
            errors = serializer.errors
            logger.warning(
                f"Submission validation failed: admission {admission.application_number}, "
                f"errors={errors}"
            )
            return Response(
                {"error": "Submission validation failed", "details": errors},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            admission = Admission.objects.select_for_update().get(pk=admission.pk)
            admission.submit()
            admission.save()
            
            logger.info(
                f"Admission submitted successfully: {admission.application_number}, "
                f"state={admission.state}"
            )
        
        self._send_whatsapp_notification(admission, 'submitted')
        
        return Response({
            "status": "submitted",
            "application_number": admission.application_number,
            "state": admission.state
        })
    
    def _send_whatsapp_notification(self, admission, notification_type):
        """SECTION 6: Send WhatsApp notification on state changes"""
        config = WhatsAppConfig.get_active_config()
        if not config or not config.notify_on_submission:
            return
        
        if notification_type == 'submitted':
            template = config.success_message_template
        else:
            return
        
        data = {
            'student_name': admission.name,
            'program_name': admission.program.name if admission.program else '',
            'application_number': admission.application_number,
        }
        
        try:
            message = template.format(**data)
            logger.info(f"WhatsApp notification prepared for {admission.application_number}: {notification_type}")
        except KeyError as e:
            logger.warning(f"WhatsApp template error: {e}")
    
    @action(detail=True, methods=['get'])
    def status(self, request, pk=None):
        """Get current status including completed steps"""
        admission = self.get_object()
        
        return Response({
            'application_number': admission.application_number,
            'state': admission.state,
            'current_step': admission.current_step,
            'completed_steps': admission.completed_steps,
            'is_submitted': admission.state != AdmissionState.DRAFT
        })
    
    @action(detail=True, methods=['post'])
    def add_note(self, request, pk=None):
        """SECTION 7: Add internal note to admission"""
        from .serializers import InternalNoteSerializer
        
        admission = self.get_object()
        serializer = InternalNoteSerializer(data={
            **request.data,
            'admission': admission.id
        })
        serializer.is_valid(raise_exception=True)
        
        author = request.user.username if request.user.is_authenticated else 'staff'
        serializer.save(author=author)
        
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
    @action(detail=True, methods=['post'])
    def transition(self, request, pk=None):
        """Admin: Transition state (approve/reject)"""
        from .serializers import StateTransitionSerializer, AdmissionDetailSerializer
        
        admission = self.get_object()
        serializer = StateTransitionSerializer(
            admission,
            data=request.data,
            context={'request': request}
        )
        serializer.is_valid(raise_exception=True)
        
        admin_user = request.user.username if request.user.is_authenticated else 'admin'
        old_state = admission.state
        admission = serializer.transition(admin_user=admin_user)
        
        if admission.state == AdmissionState.APPROVED:
            self._send_whatsapp_notification(admission, 'approved')
        elif admission.state == AdmissionState.REJECTED:
            self._send_whatsapp_notification(admission, 'rejected')
        
        cache.delete('admission_dashboard_stats')
        
        output_serializer = AdmissionDetailSerializer(admission)
        return Response(output_serializer.data)
    
    @action(detail=True, methods=['get'])
    def duplicates(self, request, pk=None):
        """SECTION 8: Check for potential duplicates"""
        admission = self.get_object()
        
        duplicates = self._check_duplicates(admission)
        
        return Response({
            'is_possible_duplicate': bool(duplicates),
            'duplicates': duplicates or []
        })
    
    @action(detail=False, methods=['post'])
    def bulk_action(self, request, pk=None):
        """SECTION 9: Bulk operations - approve, reject, export"""
        action_type = request.data.get('action')
        admission_ids = request.data.get('admission_ids', [])
        
        if not admission_ids:
            return Response(
                {'error': 'No admission IDs provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if action_type not in ['approve', 'reject', 'export']:
            return Response(
                {'error': 'Invalid action. Use: approve, reject, or export'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        admissions = Admission.objects.filter(id__in=admission_ids).select_related('program')
        
        results = {
            'total': admissions.count(),
            'success': 0,
            'failed': 0,
            'errors': []
        }
        
        with transaction.atomic():
            for admission in admissions:
                try:
                    if action_type == 'approve':
                        if admission.can_transition_to(AdmissionState.APPROVED):
                            admission.approve()
                            results['success'] += 1
                        else:
                            results['failed'] += 1
                            results['errors'].append(f"{admission.application_number}: Cannot approve")
                    
                    elif action_type == 'reject':
                        if admission.can_transition_to(AdmissionState.REJECTED):
                            reason = request.data.get('reason', 'Bulk rejection')
                            admission.reject(reason)
                            results['success'] += 1
                        else:
                            results['failed'] += 1
                            results['errors'].append(f"{admission.application_number}: Cannot reject")
                
                except Exception as e:
                    results['failed'] += 1
                    results['errors'].append(f"{admission.application_number}: {str(e)}")
        
        cache.delete('admission_dashboard_stats')
        
        return Response(results)


# ==============================================================================
# SECTION 4: INTERVIEW/VIVA PDF SHEET
# ==============================================================================

class VivaSheetView(views.APIView):
    """Generate Interview/Viva PDF Sheet"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request, pk=None):
        if not WEASYPRINT_AVAILABLE:
            return Response(
                {'error': 'WeasyPrint is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        try:
            admission = Admission.objects.select_related('program').get(pk=pk)
        except Admission.DoesNotExist:
            return Response(
                {'error': 'Admission not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        html_content = self._generate_html(admission)
        
        pdf_file = HTML(string=html_content).write_pdf()
        
        response = Response(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="viva-sheet-{admission.application_number}.pdf"'
        
        return response
    
    def _generate_html(self, admission):
        """Generate clean, professional HTML for PDF"""
        
        age = admission.age_at_submission
        if not age and admission.dob:
            today = timezone.now().date()
            age = today.year - admission.dob.year
            if (today.month, today.day) < (admission.dob.month, admission.dob.day):
                age -= 1
        
        languages = ', '.join(admission.languages_known) if admission.languages_known else 'N/A'
        if admission.languages_other:
            languages += f", {admission.languages_other}"
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <style>
                @page {{ size: A4; margin: 20mm; }}
                body {{ font-family: Arial, sans-serif; font-size: 11pt; line-height: 1.4; color: #333; }}
                .header {{ text-align: center; margin-bottom: 20px; border-bottom: 2px solid #1a5f7a; padding-bottom: 15px; }}
                .header h1 {{ margin: 0; color: #1a5f7a; font-size: 18pt; }}
                .header h2 {{ margin: 5px 0 0; color: #666; font-size: 12pt; font-weight: normal; }}
                .app-info {{ background: #f5f5f5; padding: 10px; margin-bottom: 20px; border-radius: 4px; }}
                .app-info table {{ width: 100%; }}
                .app-info td {{ padding: 3px 10px; }}
                .app-info td:first-child {{ font-weight: bold; width: 40%; }}
                .section {{ margin-bottom: 15px; }}
                .section-title {{ background: #1a5f7a; color: white; padding: 5px 10px; font-size: 11pt; margin-bottom: 8px; }}
                .field-table {{ width: 100%; }}
                .field-table td {{ padding: 4px 0; border-bottom: 1px solid #eee; }}
                .field-table td:first-child {{ width: 35%; color: #666; }}
                .rating-grid {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 10px; margin: 10px 0; }}
                .rating-item {{ border: 1px solid #ccc; padding: 8px; border-radius: 4px; }}
                .rating-label {{ font-weight: bold; margin-bottom: 5px; }}
                .rating-box {{ border-bottom: 1px solid #333; height: 20px; margin-top: 5px; }}
                .notes-section {{ border: 1px solid #ccc; min-height: 60px; margin: 10px 0; }}
                .decision {{ text-align: center; padding: 15px; background: #f9f9f9; margin: 15px 0; }}
                .decision-options {{ display: flex; justify-content: space-around; margin-top: 10px; }}
                .decision-box {{ border: 2px solid #333; padding: 10px 30px; border-radius: 4px; }}
                .signatures {{ display: flex; justify-content: space-between; margin-top: 30px; }}
                .signature-box {{ text-align: center; width: 45%; }}
                .signature-line {{ border-top: 1px solid #333; margin-top: 40px; padding-top: 5px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h1>Zainussunna Academy</h1>
                <h2>Student Admission Interview Sheet</h2>
            </div>
            
            <div class="app-info">
                <table>
                    <tr>
                        <td>Application Number:</td>
                        <td><strong>{admission.application_number}</strong></td>
                    </tr>
                    <tr>
                        <td>Program:</td>
                        <td><strong>{admission.program.name if admission.program else 'N/A'}</strong></td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <div class="section-title">Student Details</div>
                <table class="field-table">
                    <tr>
                        <td>Name:</td>
                        <td>{admission.name}</td>
                    </tr>
                    <tr>
                        <td>Date of Birth:</td>
                        <td>{admission.dob.strftime('%d-%m-%Y') if admission.dob else 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Age:</td>
                        <td>{age} years</td>
                    </tr>
                    <tr>
                        <td>Phone:</td>
                        <td>{admission.phone_country_code} {admission.phone}</td>
                    </tr>
                    <tr>
                        <td>District:</td>
                        <td>{admission.address_district}</td>
                    </tr>
                    <tr>
                        <td>Full Address:</td>
                        <td>{admission.address_house_name}, {admission.address_place}, {admission.address_post_office} - {admission.address_pin_code}, {admission.address_state}</td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <div class="section-title">Academic Background</div>
                <table class="field-table">
                    <tr>
                        <td>Madrassa Name:</td>
                        <td>{admission.madrassa_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Class Stopped:</td>
                        <td>{admission.class_stopped or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>School/College:</td>
                        <td>{admission.school_college or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Current Class:</td>
                        <td>{admission.standard or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Languages Known:</td>
                        <td>{languages}</td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <div class="section-title">Guardian Details</div>
                <table class="field-table">
                    <tr>
                        <td>Guardian Name:</td>
                        <td>{admission.guardian_name or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Relation:</td>
                        <td>{admission.guardian_relation or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Phone:</td>
                        <td>{admission.guardian_phone_country_code} {admission.guardian_phone or 'N/A'}</td>
                    </tr>
                    <tr>
                        <td>Occupation:</td>
                        <td>{admission.guardian_occupation or 'N/A'}</td>
                    </tr>
                </table>
            </div>
            
            <div class="section">
                <div class="section-title">Interview Evaluation</div>
                <div class="rating-grid">
                    <div class="rating-item">
                        <div class="rating-label">Arabic Reading</div>
                        <div class="rating-box"></div>
                    </div>
                    <div class="rating-item">
                        <div class="rating-label">Quran Recitation</div>
                        <div class="rating-box"></div>
                    </div>
                    <div class="rating-item">
                        <div class="rating-label">General Knowledge</div>
                        <div class="rating-box"></div>
                    </div>
                    <div class="rating-item">
                        <div class="rating-label">Character & Behavior</div>
                        <div class="rating-box"></div>
                    </div>
                </div>
            </div>
            
            <div class="section">
                <div class="section-title">Interview Notes</div>
                <div class="notes-section"></div>
            </div>
            
            <div class="decision">
                <strong>Decision</strong>
                <div class="decision-options">
                    <div class="decision-box">Recommended</div>
                    <div class="decision-box">Waitlist</div>
                    <div class="decision-box">Reject</div>
                </div>
            </div>
            
            <div class="signatures">
                <div class="signature-box">
                    <div class="signature-line">Interviewer Signature</div>
                </div>
                <div class="signature-box">
                    <div class="signature-line">Date</div>
                </div>
            </div>
        </body>
        </html>
        """
        return html


# ==============================================================================
# SECTION 5: FULL DATA EXPORT (EXCEL)
# ==============================================================================

class FullExportView(views.APIView):
    """Full Data Export to Excel with all fields"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'error': 'openpyxl is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        program_slug = request.query_params.get('program')
        state = request.query_params.get('state')
        search = request.query_params.get('search')
        
        queryset = Admission.objects.select_related('program').all()
        
        if program_slug:
            queryset = queryset.filter(program__slug=program_slug)
        
        if state:
            queryset = queryset.filter(state=state)
        
        if search:
            queryset = queryset.filter(
                Q(name__icontains=search) |
                Q(application_number__icontains=search)
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admissions Full Export"
        
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="1a5f7a", end_color="1a5f7a", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        headers = [
            'Application No', 'Program', 'Student Name', 'DOB', 'Age', 'Phone', 'Email',
            'District', 'State', 'Madrassa Name', 'Class Stopped', 'School', 'Languages Known',
            'Guardian Name', 'Guardian Phone', 'Guardian Occupation', 'Submitted Date',
            'Application Status', 'Current Step', 'Created At'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        for row_num, admission in enumerate(queryset, 2):
            languages = ', '.join(admission.languages_known) if admission.languages_known else ''
            if admission.languages_other:
                languages += f", {admission.languages_other}"
            
            submitted_date = admission.submitted_at.strftime('%Y-%m-%d %H:%M') if admission.submitted_at else 'N/A'
            
            row_data = [
                admission.application_number,
                admission.program.name if admission.program else '',
                admission.name,
                admission.dob.strftime('%Y-%m-%d') if admission.dob else '',
                admission.age_at_submission or '',
                f"{admission.phone_country_code} {admission.phone}",
                admission.email,
                admission.address_district,
                admission.address_state,
                admission.madrassa_name or '',
                admission.class_stopped or '',
                admission.school_college or '',
                languages,
                admission.guardian_name or '',
                f"{admission.guardian_phone_country_code} {admission.guardian_phone}" if admission.guardian_phone else '',
                admission.guardian_occupation or '',
                submitted_date,
                admission.get_state_display(),
                f"Step {admission.current_step}",
                admission.created_at.strftime('%Y-%m-%d %H:%M'),
            ]
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.alignment = cell_alignment
                cell.border = thin_border
        
        column_widths = [18, 20, 25, 12, 8, 18, 30, 18, 15, 25, 15, 25, 25, 20, 18, 20, 18, 15, 12, 18]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"admissions_full_export_{date_str}.xlsx"
        
        return Response(
            buffer.getvalue(),
            status=status.HTTP_200_OK,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )


# ==============================================================================
# ADDITIONAL VIEWS
# ==============================================================================

class InternalNoteViewSet(viewsets.ModelViewSet):
    """Internal notes for admissions - staff only"""
    serializer_class = None  # Will be set from serializers.py
    queryset = InternalNote.objects.all()
    
    def get_serializer_class(self):
        from .serializers import InternalNoteSerializer
        return InternalNoteSerializer
    
    def get_queryset(self):
        admission_id = self.kwargs.get('admission_pk')
        if admission_id:
            return InternalNote.objects.filter(admission_id=admission_id)
        return InternalNote.objects.all()


class ContentPageViewSet(viewsets.ReadOnlyModelViewSet):
    """Content pages API - read only"""
    serializer_class = None
    queryset = ContentPage.objects.filter(is_published=True)
    lookup_field = 'slug'
    
    def get_serializer_class(self):
        from .serializers import ContentPageSerializer
        return ContentPageSerializer
    
    def get_queryset(self):
        now = timezone.now()
        queryset = super().get_queryset()
        queryset = queryset.filter(
            Q(visible_from__isnull=True) | Q(visible_from__lte=now)
        ).filter(
            Q(visible_until__isnull=True) | Q(visible_until__gte=now)
        )
        return queryset


class AchievementViewSet(viewsets.ReadOnlyModelViewSet):
    """Achievements API - read only"""
    serializer_class = None
    queryset = Achievement.objects.filter(is_visible=True)
    lookup_field = 'pk'
    
    def get_serializer_class(self):
        from .serializers import AchievementSerializer
        return AchievementSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        date_from = self.request.query_params.get('from')
        date_to = self.request.query_params.get('to')
        if date_from:
            queryset = queryset.filter(date__gte=date_from)
        if date_to:
            queryset = queryset.filter(date__lte=date_to)
        return queryset


class GalleryViewSet(viewsets.ReadOnlyModelViewSet):
    """Gallery API - read only"""
    serializer_class = None
    queryset = GalleryItem.objects.filter(is_visible=True)
    lookup_field = 'pk'
    
    def get_serializer_class(self):
        from .serializers import GalleryItemSerializer
        return GalleryItemSerializer
    
    def get_queryset(self):
        return super().get_queryset().order_by('display_order')
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        items = self.get_queryset()[:10]
        serializer = self.get_serializer(items, many=True)
        return Response(serializer.data)


class EnquiryViewSet(viewsets.ModelViewSet):
    """Enquiries API"""
    serializer_class = None
    queryset = Enquiry.objects.all()
    
    def get_serializer_class(self):
        from .serializers import EnquirySerializer
        return EnquirySerializer
    
    def get_queryset(self):
        status_filter = self.request.query_params.get('status')
        if status_filter:
            self.queryset = self.queryset.filter(status=status_filter)
        
        program = self.request.query_params.get('program')
        if program:
            self.queryset = self.queryset.filter(
                Q(program_interest__slug=program) |
                Q(tagged_programs__contains=[program])
            )
        
        return self.queryset.order_by('-created_at')
    
    def create(self, request, *args, **kwargs):
        tagged = request.data.get('tagged_programs', [])
        program_interest = request.data.get('program_interest')
        
        if program_interest and str(program_interest) not in tagged:
            tagged.append(str(program_interest))
            request.data['tagged_programs'] = tagged
        
        return super().create(request, *args, **kwargs)
    
    @action(detail=True, methods=['post'])
    def update_status(self, request, pk=None):
        from .serializers import EnquiryStatusSerializer, EnquirySerializer
        
        enquiry = self.get_object()
        serializer = EnquiryStatusSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        enquiry.status = serializer.validated_data['status']
        enquiry.assigned_to = serializer.validated_data.get('assigned_to', '')
        enquiry.follow_up_notes = serializer.validated_data.get('follow_up_notes', '')
        
        if enquiry.status == 'closed':
            enquiry.closed_at = timezone.now()
        
        enquiry.save()
        
        return Response(EnquirySerializer(enquiry).data)


class AnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """Analytics API - read only"""
    serializer_class = None
    
    @action(detail=False, methods=['get'])
    def admissions(self, request):
        days = int(request.query_params.get('days', 30))
        since = timezone.now() - timedelta(days=days)
        
        admissions = Admission.objects.filter(created_at__gte=since)
        
        state_dist = dict(admissions.values('state').annotate(
            count=Count('id')
        ).values_list('state', 'count'))
        
        program_dist = dict(admissions.filter(
            program__isnull=False
        ).values('program__name').annotate(
            count=Count('id')
        ).values_list('program__name', 'count'))
        
        time_per_step = {}
        for step in ['1', '2', '3']:
            avg_time = admissions.filter(
                time_spent_per_step__has_key=step
            ).aggregate(
                avg=Avg(f'time_spent_per_step__{step}')
            )['avg']
            time_per_step[f'step_{step}'] = avg_time
        
        drop_off = {}
        for step in [1, 2, 3]:
            completed = admissions.filter(completed_steps__contains=[step]).count()
            started = admissions.filter(current_step__gte=step).count()
            if started > 0:
                drop_off[f'step_{step}'] = {
                    'started': started,
                    'completed': completed,
                    'drop_off_rate': (started - completed) / started * 100
                }
        
        return Response({
            'period_days': days,
            'total_admissions': admissions.count(),
            'state_distribution': state_dist,
            'program_distribution': program_dist,
            'avg_time_per_step': time_per_step,
            'drop_off_analysis': drop_off
        })
    
    @action(detail=False, methods=['get'])
    def dashboard(self, request):
        today = timezone.now().date()
        today_start = timezone.make_aware(datetime.combine(today, datetime.min.time()))
        
        today_count = Admission.objects.filter(created_at__gte=today_start).count()
        pending = Admission.objects.filter(state=AdmissionState.SUBMITTED).count()
        
        week_ago = today - timedelta(days=7)
        recent = Admission.objects.filter(submitted_at__gte=week_ago).count()
        
        status_breakdown = dict(Admission.objects.values('state').annotate(
            count=Count('id')
        ).values_list('state', 'count'))
        
        program_demand = dict(Admission.objects.filter(
            program__isnull=False
        ).values('program__name').annotate(
            count=Count('id')
        ).order_by('-count').values_list('program__name', 'count')[:5])
        
        return Response({
            'today': today_count,
            'pending_review': pending,
            'recent_submissions': recent,
            'status_breakdown': status_breakdown,
            'top_programs': program_demand
        })


class HealthCheckView(views.APIView):
    """Health check endpoint"""
    
    def get(self, request):
        return Response({
            'status': 'healthy',
            'timestamp': timezone.now(),
            'version': '1.0.0'
        })


class FacultyViewSet(viewsets.ReadOnlyModelViewSet):
    """Faculty API - read only"""
    serializer_class = None
    queryset = Faculty.objects.filter(is_active=True)
    lookup_field = 'pk'
    
    def get_serializer_class(self):
        from .serializers import FacultySerializer
        return FacultySerializer
    
    def get_queryset(self):
        return super().get_queryset().order_by('display_order')
    
    @action(detail=False, methods=['get'])
    def latest(self, request):
        members = self.get_queryset()[:10]
        serializer = self.get_serializer(members, many=True)
        return Response(serializer.data)


class AdmissionExportView(views.APIView):
    """Legacy export endpoint"""
    
    def get(self, request):
        if not OPENPYXL_AVAILABLE:
            return Response(
                {'error': 'openpyxl is not installed'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        program_slug = request.query_params.get('program')
        status_filter = request.query_params.get('status')
        
        queryset = Admission.objects.select_related('program')
        
        if program_slug:
            queryset = queryset.filter(program__slug=program_slug)
        
        if status_filter:
            queryset = queryset.filter(state=status_filter)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Admissions"
        
        header_font_white = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        headers = ['Student Name', 'Age', 'Phone Number', 'Full Address', 'Parent / Guardian Name', 'Program Name', 'Admission Status']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        for row_num, admission in enumerate(queryset, 2):
            full_address = f"{admission.address_house_name}, {admission.address_place}, {admission.address_post_office} - {admission.address_pin_code}, {admission.address_district}, {admission.address_state}"
            
            row_data = [
                admission.name,
                admission.age_at_submission or admission.age,
                f"{admission.phone_country_code} {admission.phone}",
                full_address,
                admission.guardian_name,
                admission.program.name if admission.program else '',
                admission.get_state_display(),
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)
                ws.cell(row=row_num, column=col).alignment = Alignment(horizontal='left')
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        date_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        program_part = program_slug if program_slug else 'all'
        filename = f"admissions_{program_part}_{date_str}.xlsx"
        
        return Response(
            buffer.getvalue(),
            status=status.HTTP_200_OK,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )


class WhatsAppConfigViewSet(viewsets.ModelViewSet):
    """WhatsApp Configuration API."""
    serializer_class = None
    queryset = WhatsAppConfig.objects.all()
    
    def get_serializer_class(self):
        from .serializers import WhatsAppConfigSerializer
        return WhatsAppConfigSerializer
    
    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'generate_message', 'active']:
            return []
        return [IsAuthenticated()]
    
    def get_queryset(self):
        if self.action in ['list', 'retrieve']:
            return WhatsAppConfig.objects.filter(is_active=True)
        return WhatsAppConfig.objects.all()
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        config = WhatsAppConfig.get_active_config()
        if config:
            serializer = self.get_serializer(config)
            return Response(serializer.data)
        return Response({'error': 'No active WhatsApp configuration'}, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['post'])
    def generate_message(self, request):
        admission_id = request.data.get('admission_id')
        message_type = request.data.get('message_type', 'success')
        
        if not admission_id:
            return Response({'error': 'admission_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            admission = Admission.objects.get(id=admission_id)
        except Admission.DoesNotExist:
            return Response({'error': 'Admission not found'}, status=status.HTTP_404_NOT_FOUND)
        
        config = WhatsAppConfig.get_active_config()
        if not config:
            return Response({'error': 'WhatsApp not configured'}, status=status.HTTP_400_BAD_REQUEST)
        
        data = {
            'student_name': admission.name,
            'program_name': admission.program.name if admission.program else '',
            'standard': admission.standard or '',
            'phone': f"{admission.phone_country_code} {admission.phone}",
            'guardian_name': admission.guardian_name or '',
            'guardian_relation': admission.guardian_relation or '',
            'guardian_phone': f"{admission.guardian_phone_country_code} {admission.guardian_phone}" if admission.guardian_phone else '',
            'application_number': admission.application_number,
        }
        
        if message_type == 'success':
            message = config.format_success_message(data)
        else:
            message = config.format_admission_message(data)
        
        phone = config.phone_number.replace('+', '').replace(' ', '')
        whatsapp_url = f"https://wa.me/{phone}?text={urllib.parse.quote(message)}"
        
        return Response({
            'message': message,
            'whatsapp_url': whatsapp_url,
            'phone_number': config.phone_number
        })


class ProgramViewSet(viewsets.ModelViewSet):
    """Program API - Dynamic schema generation for frontend."""
    serializer_class = None
    queryset = Program.objects.filter(is_active=True)
    
    def get_serializer_class(self):
        from .serializers import ProgramSerializer, ProgramSummarySerializer, ProgramFieldSerializer
        if self.action == 'list':
            return ProgramSummarySerializer
        return ProgramSerializer
    
    @action(detail=True, methods=['get'])
    def schema(self, request, pk=None):
        from .serializers import ProgramFieldSerializer, ProgramSummarySerializer
        
        program = self.get_object()
        fields = program.fields.filter(is_visible=True).order_by('step', 'display_order')
        field_serializer = ProgramFieldSerializer(fields, many=True)
        
        return Response({
            'program': ProgramSummarySerializer(program).data,
            'steps': {
                step: list(fields.filter(step=step).values())
                for step in fields.values_list('step', flat=True).distinct()
            },
            'schema': field_serializer.data,
            'age_range': {'min': program.min_age, 'max': program.max_age}
        })

